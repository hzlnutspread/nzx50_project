import time
import datetime
import openpyxl
import pandas as pd
from paramiko import *

import myconstants

# list of companies
tickers = ['fph', 'aia', 'spk', 'mft', 'cen', 'ift', 'mel', 'fbu', 'rym', 'ebo', 'atm', 'mcy', 'cnu', 'sum', 'gmt',
           'skc', 'pot', 'fre', 'pct', 'kpg', 'whs', 'gne', 'pfi', 'arg', 'pph', 'hgh', 'vhp', 'skl', 'arv', 'vct',
           'kmd', 'spg', 'oca', 'peb', 'air', 'scl', 'ipl', 'sko', 'vgl', 'mnw', 'nzx', 'rbd', 'san', 'fsf', 'thl',
           'sml', 'skt', 'erd']
period1 = int(time.mktime(datetime.datetime(2016, 9, 30, 11, 59).timetuple()))
period2 = int(time.mktime(datetime.datetime.today().timetuple()))
interval = '1d'

# connect to SFTP server
host = 'nfs.interest.co.nz'
transport = Transport(host)
transport.connect(None, username=myconstants.USERNAME, password=myconstants.PASSWORD)
sftp = SFTPClient.from_transport(transport)

# 1. create share price csv
for ticker in tickers:
    print(f'retrieving {ticker} share price')
    query_string = f'https://query1.finance.yahoo.com/v7/finance/download/{ticker}.NZ?period1={period1}&period2={period2}' \
                   f'&interval={interval}&events=history&includeAdjustedClose=true'

    df = pd.read_csv(query_string)

    arr = []
    for index, row in df.iterrows():
        date = pd.to_datetime(row['Date'])
        if date.weekday() == 4:
            subarr = [date.strftime('%d-%b-%Y'), row['Close']]
            arr.append(subarr)

    df2 = pd.DataFrame(arr, columns=['Date', 'Close'])
    df2.to_csv(f'//SERVER/jdjl/interest-nz/interest.co.nz/chart_data/investing/{ticker}-share-price.csv', index=False,
               header=False)

    # upload the files to server
    local_shareprice = f'K:\\interest-nz\\interest.co.nz\\chart_data\\investing\\{ticker}-share-price.csv'
    remote_shareprice = f'/var/www/drupal8.interest.co.nz/web/sites/default/files/charts-csv/chart_data/investing/{ticker}-share-price.csv'
    sftp.put(local_shareprice, remote_shareprice)

# 2. create capitalisation csv
wb = openpyxl.load_workbook('C:/Users/User/Desktop/NZX 50.xlsx', read_only=True, data_only=True)
for ticker in tickers:
    print(f'retrieving {ticker} capitalisation')

    arr = []
    for sheet in wb.sheetnames:
        sh = wb[f'{sheet}']

        checker = False
        for row in sh.iter_rows():
            if checker is True:
                break
            for cell in row:

                if cell.value == f'{ticker}'.upper():
                    checker = True
                    date_val = sh.cell(row=cell.row, column=12).value
                    if date_val is None:
                        raise Exception('You fucked up ' + sheet)
                    cap_val = int((sh.cell(row=cell.row, column=8).value) / 1_000_000)
                    subarr = [date_val.strftime('%d-%b-%Y'), cap_val]
                    arr.append(subarr)
                    break

    df = pd.DataFrame(arr, columns=['Date', 'Capitalisation'])
    df.to_csv(f'//SERVER/jdjl/interest-nz/interest.co.nz/chart_data/investing/{ticker}-cap.csv', header=False,
              index=False)

    # upload the files to server
    local_cap = f'K:\\interest-nz\\interest.co.nz\\chart_data\\investing\\{ticker}-cap.csv'
    remote_cap = f'/var/www/drupal8.interest.co.nz/web/sites/default/files/charts-csv/chart_data/investing/{ticker}-cap.csv'
    sftp.put(local_cap, remote_cap)

# 3. create cap $ change csv
for ticker in tickers:
    print(f'creating {ticker} cap $ change')

    cap_csv = f'//SERVER/jdjl/interest-nz/interest.co.nz/chart_data/investing/{ticker}-cap.csv'
    df = pd.read_csv(cap_csv)

    df.columns = ['Date', 'Cap']

    cap_change = pd.Series(df['Cap'].diff()).array  # do this with a loop
    date = pd.Series(df['Date']).array

    df2 = pd.DataFrame(date, columns=['Date'])
    df3 = pd.DataFrame(cap_change, columns=['Cap'])
    df4 = pd.concat([df2, df3], axis=1)
    df5 = df4.drop(index=0)
    df5.to_csv(f'//SERVER/jdjl/interest-nz/interest.co.nz/chart_data/investing/{ticker}-cap-$change.csv', index=False,
               header=False)

    # upload the files to server
    local_capabschange = f'K:\\interest-nz\\interest.co.nz\\chart_data\\investing\\{ticker}-cap-$change.csv'
    remote_capabschange = f'/var/www/drupal8.interest.co.nz/web/sites/default/files/charts-csv/chart_data/investing/{ticker}-cap-$change.csv'
    sftp.put(local_capabschange, remote_capabschange)

# 4. create cap % change csv
for ticker in tickers:
    print(f'creating {ticker} cap % change')

    cap_csv = f'//SERVER/jdjl/interest-nz/interest.co.nz/chart_data/investing/{ticker}-cap.csv'
    df = pd.read_csv(cap_csv)

    df.columns = ['Date', 'Cap']

    cap_change = pd.Series(df['Cap'].pct_change() * 100).round(decimals=2).array  # do this with a loop
    date = pd.Series(df['Date']).array

    df2 = pd.DataFrame(date, columns=['Date'])
    df3 = pd.DataFrame(cap_change, columns=['Cap'])
    df4 = pd.concat([df2, df3], axis=1)
    df5 = df4.drop(index=0)
    df5.to_csv(f'//SERVER/jdjl/interest-nz/interest.co.nz/chart_data/investing/{ticker}-cap-change.csv', index=False,
               header=False)

    # upload the files to server
    local_capchange = f'K:\\interest-nz\\interest.co.nz\\chart_data\\investing\\{ticker}-cap-change.csv'
    remote_capchange = f'/var/www/drupal8.interest.co.nz/web/sites/default/files/charts-csv/chart_data/investing/{ticker}-cap-change.csv'
    sftp.put(local_capchange, remote_capchange)

# 5. create rank csv
for ticker in tickers:
    print(f'retrieving {ticker} rank')

    arr = []
    for sheet in wb.sheetnames:
        sh = wb[f'{sheet}']

        checker = False
        for row in sh.iter_rows():
            if checker is True:
                break
            for cell in row:

                if cell.value == f'{ticker}'.upper():
                    checker = True
                    date_val = sh.cell(row=cell.row, column=12).value
                    if date_val is None:
                        raise Exception('You fucked up ' + sheet)
                    cell_val = sh.cell(row=cell.row, column=4).value
                    subarr = [date_val.strftime('%d-%b-%Y'), cell_val]
                    arr.append(subarr)
                    break

    df = pd.DataFrame(arr, columns=['Date', 'Rank'])
    df.to_csv(f'//SERVER/jdjl/interest-nz/interest.co.nz/chart_data/investing/{ticker}-rank.csv', index=False,
              header=False)

    # upload the files to server
    local_rank = f'K:\\interest-nz\\interest.co.nz\\chart_data\\investing\\{ticker}-rank.csv'
    remote_rank = f'/var/www/drupal8.interest.co.nz/web/sites/default/files/charts-csv/chart_data/investing/{ticker}-rank.csv'
    sftp.put(local_rank, remote_rank)

print("Success: Program has finished updating and uploading")