import openpyxl
import pandas as pd

tickers = ['fph', 'aia', 'spk', 'mft', 'cen', 'ift', 'fbu', 'mel', 'rym', 'ebo', 'atm', 'mcy', 'cnu', 'sum', 'gmt',
           'skc', 'pot', 'fre', 'pct', 'kpg', 'zel', 'gne', 'pfi', 'arg', 'pph', 'hgh', 'vhp', 'skl', 'arv', 'vct',
           'kmd', 'peb', 'spg', 'oca', 'air', 'scl', 'sko', 'ipl', 'vgl', 'nzx', 'tpw', 'rbd', 'skt', 'fsf', 'san',
           'thl', 'sml', 'nph']

wb = openpyxl.load_workbook('C:/Users/ken/Desktop/NZX 50.xlsx', read_only=True, data_only=True)

# create the cap csvs
for ticker in tickers:
    print(f'retrieving {ticker} capitalisation')

    arr = []
    for sheet in wb.sheetnames:
        sh = wb[f'{sheet}']

        checker = False
        for row in sh.iter_rows():
            if checker is True:
                break
            for cell in row:

                if cell.value == f'{ticker}'.upper():
                    checker = True
                    date_val = sh.cell(row=cell.row, column=12).value
                    if date_val is None:
                        raise Exception('You fucked up ' + sheet)
                    cap_val = int((sh.cell(row=cell.row, column=8).value) / 1_000_000)
                    subarr = [date_val.strftime('%d-%b-%Y'), cap_val]
                    arr.append(subarr)
                    break

    df = pd.DataFrame(arr, columns=['Date', 'Capitalisation'])
    df.to_csv(f'./csvs/{ticker}-cap.csv', header=False, index=False)

# create the cap change% csvs
for ticker in tickers:
    print(f'opening {ticker}-cap')

    cap_csv = f'./csvs/{ticker}-cap.csv'
    df = pd.read_csv(cap_csv)

    df.columns = ['Date', 'Cap']

    cap_change = pd.Series(df['Cap'].pct_change() * 100).round(decimals=2).array
    date = pd.Series(df['Date']).array

    df2 = pd.DataFrame(date, columns=['Date'])
    df3 = pd.DataFrame(cap_change, columns=['Cap'])
    df4 = pd.concat([df2, df3], axis=1)
    df5 = df4.drop(index=0)

    df5.to_csv(f'./csvs/{ticker}-cap-change%.csv', index=False, header=False)

# create the cap change$ csvs
for ticker in tickers:
    print(f'creating {ticker} cap $ change')

    cap_csv = f'./csvs/{ticker}-cap.csv'
    df = pd.read_csv(cap_csv)

    df.columns = ['Date', 'Cap']

    cap_change = pd.Series(df['Cap'].diff()).array
    date = pd.Series(df['Date']).array

    df2 = pd.DataFrame(date, columns=['Date'])
    df3 = pd.DataFrame(cap_change, columns=['Cap'])
    df4 = pd.concat([df2, df3], axis=1)
    df5 = df4.drop(index=0)

    df5.to_csv(f'./csvs/{ticker}-cap-$change.csv', index=False, header=False)
