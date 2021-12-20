import openpyxl
import pandas as pd

tickers = ['fph', 'aia', 'spk', 'mft', 'cen', 'ift', 'fbu', 'mel', 'rym', 'ebo', 'atm', 'mcy', 'cnu', 'sum', 'gmt',
           'skc', 'pot', 'fre', 'pct', 'kpg', 'zel', 'gne', 'pfi', 'arg', 'pph', 'hgh', 'vhp', 'skl', 'arv', 'vct',
           'kmd', 'peb', 'spg', 'oca', 'air', 'scl', 'sko', 'ipl', 'vgl', 'nzx', 'tpw', 'rbd', 'skt', 'fsf', 'san',
           'thl', 'sml', 'nph']
sheets = ['NZX50 smartshares 30 Sep 2020', '09 Oct 2020', '16 Oct 2020', '23 Oct 2020', '30 Oct 2020', '06 Nov 2020',
          '13 Nov 2020', '20 Nov 2020', '27 Nov 2020', '27 Nov 2020', '04 Dec 2020', '11 Dec 2020', '18 Dec 2020',
          '25 Dec 2020', '01 Jan 2021', '08 Jan 2021', '15 Jan 2021', '22 Jan 2021', '29 Jan 2021', '05 Feb 2021',
          '12 Feb 2021', '19 Feb 2021', '26 Feb 2021', '05 Mar 2021', '12 Mar 2021', '19 Mar 2021', '26 Mar 2021',
          '02 Apr 2021', '09 Apr 2021', '16 Apr 2021', '23 Apr 2021', '30 Apr 2021', '07 May 2021', '14 May 2021',
          '21 May 2021', '28 May 2021', '04 June 2021', '11 June 2021', '18 June 2021', '25 June 2021', '02 July 2021',
          '09 July 2021', '16 July 2021', '23 July 2021', '30 July 2021', '06 Aug 2021', '13 Aug 2021', '20 Aug 2021',
          '27 Aug 2021', '03 Sep 2021', '10 Sep 2021', '17 Sep 2021', '24 Sep 2021', '01 Oct 2021', '08 Oct 2021',
          '15 Oct 2021', '22 Oct 2021', '29 Oct 2021', '05 Nov 2021', '12 Nov 2021', '19 Nov 2021', '26 Nov 2021',
          '03 Dec 2021', '10 Dec 2021']

wb = openpyxl.load_workbook('C:/Users/ken/Desktop/NZX 50.xlsx', read_only=True)

for ticker in tickers:
    print(f'retrieving {ticker}')

    arr = []
    for sheet in sheets:
        sh = wb[f'{sheet}']

        for row in sh.iter_rows():
            for cell in row:
                if cell.value == f'{ticker}'.upper():
                    date_val = sh.cell(row=cell.row, column=12).value
                    cell_val = sh.cell(row=cell.row, column=4).value
                    subarr = [date_val.strftime('%d-%b-%Y'), cell_val]
                    arr.append(subarr)

    df = pd.DataFrame(arr, columns=['Date', 'Rank'])
    df.to_csv(f'./csvs/{ticker}-rank.csv', index=False, header=False)
